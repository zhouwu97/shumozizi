"""从当前正式结果填充竞赛提供的三个 Excel 提交模板。"""

from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "results" / "raw"
TEMPLATE_ROOT = ROOT / "problem" / "attachments" / "附件"
SUBMISSION = ROOT / "submission"


def _load(question: str) -> dict[str, object]:
    """读取一个问题的当前 JSON 结果。"""

    return json.loads((RAW / f"{question.lower()}.json").read_text(encoding="utf-8"))


def _write_action_row(
    sheet: object,
    row: int,
    action: dict[str, object],
    duration: float,
    *,
    columns: tuple[int, ...],
    missile_column: int | None = None,
) -> None:
    """按模板显式列映射写入一个动作，保留编号列和原有样式。"""

    values = [
        action["heading_deg"],
        action["speed_mps"],
        *action["release_point_m"],
        *action["burst_point_m"],
        duration,
    ]
    if len(columns) != len(values):
        raise ValueError("提交模板动作列映射长度不正确")
    for column, value in zip(columns, values, strict=True):
        cell = sheet.cell(row=row, column=column)
        cell.value = float(value)
        cell.number_format = "0.000"
    if missile_column is not None:
        sheet.cell(row=row, column=missile_column).value = action["assigned_missile"]


def _prepare(filename: str) -> tuple[object, Path]:
    """复制模板并返回工作簿与提交路径。"""

    SUBMISSION.mkdir(parents=True, exist_ok=True)
    target = SUBMISSION / filename
    shutil.copy2(TEMPLATE_ROOT / filename, target)
    return load_workbook(target), target


def _fill_result1() -> dict[str, object]:
    """填充 Q3 的 FY1 三弹结果。"""

    payload = _load("Q3")
    actions = payload["result"]["actions"]
    durations = payload["result"]["individual_durations_s"]
    workbook, target = _prepare("result1.xlsx")
    sheet = workbook.active
    for row, (action, duration) in enumerate(zip(actions, durations), start=2):
        _write_action_row(
            sheet,
            row,
            action,
            duration,
            columns=(1, 2, 4, 5, 6, 7, 8, 9, 10),
        )
    workbook.save(target)
    return {"file": "submission/result1.xlsx", "filled_actions": len(actions)}


def _fill_result2() -> dict[str, object]:
    """填充 Q4 的三架无人机单弹结果。"""

    payload = _load("Q4")
    actions = payload["result"]["actions"]
    durations = payload["result"]["individual_durations_s"]
    workbook, target = _prepare("result2.xlsx")
    sheet = workbook.active
    for row, (action, duration) in enumerate(zip(actions, durations), start=2):
        _write_action_row(
            sheet,
            row,
            action,
            duration,
            columns=(2, 3, 4, 5, 6, 7, 8, 9, 10),
        )
    workbook.save(target)
    return {"file": "submission/result2.xlsx", "filled_actions": len(actions)}


def _fill_result3() -> dict[str, object]:
    """填充 Q5 的可变动作数结果，未激活槽位保持空白。"""

    payload = _load("Q5")
    actions = payload["result"]["actions"]
    durations = payload["result"]["individual_assigned_duration_s"]
    paired = list(zip(actions, durations))
    workbook, target = _prepare("result3.xlsx")
    sheet = workbook.active
    filled_rows: list[int] = []
    for drone_index, drone in enumerate(("FY1", "FY2", "FY3", "FY4", "FY5")):
        group = sorted(
            (item for item in paired if item[0]["drone"] == drone),
            key=lambda item: item[0]["release_time_s"],
        )
        for bomb_index, (action, duration) in enumerate(group):
            row = 2 + 3 * drone_index + bomb_index
            _write_action_row(
                sheet,
                row,
                action,
                duration,
                columns=(2, 3, 5, 6, 7, 8, 9, 10, 11),
                missile_column=12,
            )
            filled_rows.append(row)
    workbook.save(target)
    return {
        "file": "submission/result3.xlsx",
        "filled_actions": len(actions),
        "filled_rows": filled_rows,
        "unused_rows": [row for row in range(2, 17) if row not in filled_rows],
    }


def main() -> int:
    """生成三个提交工作簿和结构化填充收据。"""

    tables = [_fill_result1(), _fill_result2(), _fill_result3()]
    payload = {
        "schema_version": "1.0",
        "tables": tables,
        "semantics": {
            "bomb_numbering": "per-drone chronological release order",
            "individual_duration": "single bomb exact duration for assigned missile",
            "unused_slot": "kept blank",
        },
        "metrics": {
            "result1_filled_actions": tables[0]["filled_actions"],
            "result2_filled_actions": tables[1]["filled_actions"],
            "result3_filled_actions": tables[2]["filled_actions"],
        },
    }
    output = RAW / "submission_tables.json"
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload["metrics"], ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
