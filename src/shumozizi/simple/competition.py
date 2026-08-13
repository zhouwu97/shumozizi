"""Competition-First v3.1 的路线锦标赛、实验价值和洞察产物。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from shumozizi.core.io import (
    ContractError,
    atomic_json,
    load_json,
    resolve_inside,
    sha256_file,
)
from shumozizi.simple.results import read_result_index
from shumozizi.simple.state import read_simple_state, utc_now

ROUTE_COMPETITION_PATH = Path("analysis/ROUTE_COMPETITION.md")
ROUTE_TOURNAMENT_PATH = Path("analysis/route_tournament.json")
NEXT_EXPERIMENTS_PATH = Path("analysis/NEXT_EXPERIMENTS.md")
INSIGHTS_PATH = Path("analysis/INSIGHTS.md")
ANSWER_MAP_PATH = Path("analysis/answer_map.json")


def _require_text(value: object, label: str) -> str:
    """验证非空文本字段。"""
    if not isinstance(value, str) or not value.strip():
        raise ContractError(f"{label} 必须是非空文本")
    return value.strip()


def validate_route_competition(payload: dict[str, Any]) -> list[str]:
    """校验路线竞争至少存在基线和真正不同的比较路线。

    Args:
        payload: 路线竞争的结构化输入。

    Returns:
        可读错误列表，空列表表示通过。
    """
    errors: list[str] = []
    baseline = payload.get("baseline")
    candidates = payload.get("candidates")
    if not isinstance(baseline, dict):
        return ["ROUTE_COMPETITION 必须包含 baseline"]
    try:
        baseline_structure = _require_text(baseline.get("mathematical_structure"), "baseline.mathematical_structure")
    except ContractError as exc:
        errors.append(str(exc))
        baseline_structure = ""
    if not isinstance(candidates, list) or not candidates:
        errors.append("至少需要一条与 baseline 真正不同的竞争路线或反证路线")
        return errors
    distinct = False
    for index, candidate in enumerate(candidates):
        if not isinstance(candidate, dict):
            errors.append(f"candidates[{index}] 必须是对象")
            continue
        try:
            structure = _require_text(candidate.get("mathematical_structure"), f"candidates[{index}].mathematical_structure")
            _require_text(candidate.get("probe"), f"candidates[{index}].probe")
        except ContractError as exc:
            errors.append(str(exc))
            continue
        if structure != baseline_structure:
            distinct = True
    if not distinct:
        errors.append("仅替换遗传算法、粒子群或差分进化等求解器不构成不同路线")
    return errors


def _number(value: object, label: str) -> float:
    """读取有限数值，避免比较时把字符串或 NaN 当作真实成绩。"""
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        raise ContractError(f"{label} 必须是数值")
    number = float(value)
    if number != number or number in {float("inf"), float("-inf")}:
        raise ContractError(f"{label} 必须是有限数值")
    return number


def _route_records(
    run_dir: Path, *, question_id: str, route: dict[str, Any], exact_metric: str
) -> tuple[float, float]:
    """复验一条路线的 exact、可行性、稳健性和实际耗时结果。"""
    route_id = _require_text(route.get("route_id"), "route.route_id")
    index = {
        item["result_id"]: item
        for item in read_result_index(run_dir)["results"]
    }

    def result(result_id: object, label: str) -> dict[str, Any]:
        identifier = _require_text(result_id, f"{route_id}.{label}")
        item = index.get(identifier)
        if item is None:
            raise ContractError(f"{route_id}.{label} 未绑定实际 result_id: {identifier}")
        if (
            item.get("question_id") != question_id
            or item.get("execution_mode") != "production"
            or item.get("execution_valid") is not True
        ):
            raise ContractError(f"{route_id}.{label} 必须是本问 execution_valid 的 production 结果")
        return item

    exact = result(route.get("exact_score_result_id"), "exact_score_result_id")
    exact_score = _number(exact.get("metrics", {}).get(exact_metric), f"{route_id}.{exact_metric}")
    feasible = result(route.get("feasibility_result_id"), "feasibility_result_id")
    if feasible.get("metrics", {}).get("feasible") is not True or route.get("feasible") is not True:
        raise ContractError(f"{route_id} 缺少已执行的可行性通过结果")
    robust = result(route.get("robustness_result_id"), "robustness_result_id")
    if robust.get("metrics", {}).get("robustness_passed") is not True:
        raise ContractError(f"{route_id} 缺少通过的稳健性结果")
    probes = route.get("probe_result_ids")
    budgets = route.get("budget_result_ids")
    if not isinstance(probes, list) or not probes:
        raise ContractError(f"{route_id} 必须绑定至少一个实际 probe_result_id")
    if not isinstance(budgets, list) or not budgets:
        raise ContractError(f"{route_id} 必须绑定 budget_result_ids")
    for probe_id in probes:
        result(probe_id, "probe_result_ids")
    unique_budget = list(dict.fromkeys(budgets))
    measured_runtime = sum(
        _number(result(result_id, "budget_result_ids").get("duration_seconds"), f"{route_id}.duration_seconds")
        for result_id in unique_budget
    )
    declared_runtime = _number(route.get("runtime_seconds"), f"{route_id}.runtime_seconds")
    if abs(measured_runtime - declared_runtime) > 1e-6:
        raise ContractError(f"{route_id}.runtime_seconds 必须等于绑定实际执行的耗时和")
    for field in ("mathematical_structure", "mechanism_value", "failure_mode", "switch_condition"):
        _require_text(route.get(field), f"{route_id}.{field}")
    assumptions = route.get("assumptions")
    if not isinstance(assumptions, list) or not assumptions or not all(
        isinstance(item, str) and item.strip() for item in assumptions
    ):
        raise ContractError(f"{route_id}.assumptions 必须是非空文本数组")
    return exact_score, measured_runtime


def _best_route(
    scores: dict[str, float], direction: str) -> str:
    """按统一 exact 目标返回最优路线 ID。"""
    if direction == "minimize":
        return min(scores, key=scores.get)
    return max(scores, key=scores.get)


def _improvement_ratio(baseline: float, winner: float, direction: str) -> float:
    """计算相对 baseline 的正向 exact 改善比例。"""
    denominator = max(abs(baseline), 1e-12)
    return (baseline - winner) / denominator if direction == "minimize" else (winner - baseline) / denominator


def _validate_proxy_order(
    run_dir: Path,
    routes: list[dict[str, Any]],
    *,
    direction: str,
    exact_winner: str,
) -> None:
    """当路线登记 proxy 时，拒绝 proxy 与 exact 排名反转。"""
    index = {item["result_id"]: item for item in read_result_index(run_dir)["results"]}
    proxy_scores: dict[str, float] = {}
    for route in routes:
        result = index[route["exact_score_result_id"]]
        value = result.get("metrics", {}).get("proxy_score")
        if value is not None:
            proxy_scores[route["route_id"]] = _number(value, f"{route['route_id']}.proxy_score")
    if len(proxy_scores) >= 2 and _best_route(proxy_scores, direction) != exact_winner:
        raise ContractError("proxy 与统一 exact 的路线排序反转，不能冻结路线赢家")


def _validate_route_tournament(run_dir: Path, payload: dict[str, Any]) -> tuple[dict[str, Any], str]:
    """验证真实路线锦标赛并从执行事实推导强度。"""
    if payload.get("schema_version") != "1.0":
        raise ContractError("route_tournament 必须使用 schema_version 1.0")
    state = read_simple_state(run_dir)
    question_id = _require_text(payload.get("question_id"), "question_id")
    if question_id not in state["required_questions"]:
        raise ContractError("route_tournament question_id 不是必答问题")
    comparison = payload.get("comparison")
    if not isinstance(comparison, dict):
        raise ContractError("route_tournament 缺少 comparison")
    exact_metric = _require_text(comparison.get("exact_metric"), "comparison.exact_metric")
    direction = comparison.get("objective_direction")
    if direction not in {"minimize", "maximize"}:
        raise ContractError("comparison.objective_direction 必须为 minimize 或 maximize")
    if comparison.get("budget_kind") not in {"wall_seconds"}:
        raise ContractError("当前路线锦标赛只接受可复验的 wall_seconds 共同预算")
    budget_tolerance = _number(comparison.get("budget_tolerance_ratio"), "comparison.budget_tolerance_ratio")
    improvement_threshold = _number(
        comparison.get("significant_improvement_ratio"),
        "comparison.significant_improvement_ratio",
    )
    if budget_tolerance < 0 or improvement_threshold < 0:
        raise ContractError("路线预算容差和显著改善阈值不能为负")
    baseline = payload.get("baseline")
    candidates = payload.get("candidates")
    if not isinstance(baseline, dict) or not isinstance(candidates, list):
        raise ContractError("route_tournament 必须包含 baseline 和 candidates")
    routes = [baseline, *candidates]
    if not all(isinstance(route, dict) for route in routes):
        raise ContractError("baseline 与 candidates 必须都是对象")
    exemption = payload.get("exemption")
    if exemption is not None:
        if not isinstance(exemption, dict) or exemption.get("type") not in {"simple_recompute", "analytic_solution"}:
            raise ContractError("路线豁免只能用于 simple_recompute 或 analytic_solution")
        _require_text(exemption.get("reason"), "exemption.reason")
    elif len(candidates) < 2:
        raise ContractError("核心问题至少需要两条与 baseline 数学结构不同的竞争路线")
    route_ids = [_require_text(route.get("route_id"), "route.route_id") for route in routes]
    if len(set(route_ids)) != len(route_ids):
        raise ContractError("route_id 必须唯一")
    baseline_structure = _require_text(baseline.get("mathematical_structure"), "baseline.mathematical_structure")
    candidate_structures = [
        _require_text(route.get("mathematical_structure"), f"{route['route_id']}.mathematical_structure")
        for route in candidates
    ]
    if exemption is None and (
        any(structure == baseline_structure for structure in candidate_structures)
        or len(set(candidate_structures)) < 2
    ):
        raise ContractError("竞争路线必须彼此且相对 baseline 具有不同 mathematical_structure")
    scores: dict[str, float] = {}
    runtimes: dict[str, float] = {}
    for route in routes:
        score, runtime = _route_records(run_dir, question_id=question_id, route=route, exact_metric=exact_metric)
        scores[route["route_id"]] = score
        runtimes[route["route_id"]] = runtime
    baseline_runtime = runtimes[baseline["route_id"]]
    if baseline_runtime <= 0:
        raise ContractError("baseline 的实际共同预算必须大于零")
    unfair = [
        route_id
        for route_id, runtime in runtimes.items()
        if abs(runtime - baseline_runtime) / baseline_runtime > budget_tolerance
    ]
    if unfair:
        raise ContractError("路线共同预算不公平: " + ", ".join(sorted(unfair)))
    selection = payload.get("selection")
    if not isinstance(selection, dict):
        raise ContractError("route_tournament 缺少 selection")
    winner = _require_text(selection.get("winner_route_id"), "selection.winner_route_id")
    fallback = _require_text(selection.get("fallback_route_id"), "selection.fallback_route_id")
    _require_text(selection.get("selection_rationale"), "selection.selection_rationale")
    if winner not in scores or fallback not in scores or fallback == winner:
        raise ContractError("winner 与 fallback 必须是不同的已执行路线")
    exact_winner = _best_route(scores, direction)
    if winner != exact_winner:
        raise ContractError("路线赢家必须由统一 exact scorer 的实际最优结果决定")
    _validate_proxy_order(run_dir, routes, direction=direction, exact_winner=exact_winner)
    improvement = _improvement_ratio(scores[baseline["route_id"]], scores[winner], direction)
    near_bound = bool(selection.get("baseline_near_bound") is True and winner == baseline["route_id"])
    strong = bool(
        (improvement >= improvement_threshold or near_bound)
        and (len(candidates) >= 1 or exemption is not None)
        and all(_require_text(route.get("mechanism_value"), "route.mechanism_value") for route in routes)
    )
    strength = "strong" if strong else "qualified"
    normalized = {
        "schema_version": "1.0",
        "run_id": run_dir.name,
        "question_id": question_id,
        "core_question": payload.get("core_question") is True,
        "comparison": comparison,
        "baseline": baseline,
        "candidates": candidates,
        "selection": selection,
        "exemption": exemption,
        "scores": scores,
        "measured_runtime_seconds": runtimes,
        "derived_strength": strength,
        "result_index_sha256": sha256_file(run_dir / "results" / "index.json"),
        "generated_at": utc_now(),
    }
    return normalized, strength


def write_route_tournament(run_dir: Path, payload: dict[str, Any]) -> dict[str, Any]:
    """保存绑定实际执行、统一 exact 和共同预算的路线锦标赛。

    Args:
        run_dir: 当前运行目录。
        payload: 路线及其实际执行结果引用。

    Returns:
        写入的、由系统推导强度的锦标赛对象。

    Raises:
        ContractError: 任何路线未执行、非同目标/同预算或试图以 proxy 决胜。
    """
    normalized, _ = _validate_route_tournament(run_dir, payload)
    atomic_json(run_dir / ROUTE_TOURNAMENT_PATH, normalized)
    return normalized


def require_route_tournament_for_paper(run_dir: Path) -> dict[str, Any]:
    """要求核心问题存在仍绑定当前结果的 strong 路线锦标赛。"""
    path = run_dir / ROUTE_TOURNAMENT_PATH
    if not path.is_file():
        raise ContractError("不能进入论文阶段：缺少真实路线锦标赛 analysis/route_tournament.json")
    payload = load_json(path)
    if payload.get("result_index_sha256") != sha256_file(run_dir / "results" / "index.json"):
        raise ContractError("路线锦标赛已因结果索引变化失效，必须重新比较")
    normalized, strength = _validate_route_tournament(run_dir, payload)
    if payload.get("derived_strength") != strength:
        raise ContractError("路线锦标赛声明的强度与当前实际结果不一致")
    if normalized["core_question"] and strength != "strong":
        raise ContractError("核心问题路线强度仍为 weak/qualified，必须继续实验、切换路线或进入 blocked")
    return normalized


def validate_next_experiments(payload: dict[str, Any]) -> list[str]:
    """检查实验队列是否说明能够改变的决定，而非收集装饰性结果。

    Args:
        payload: 实验队列对象。

    Returns:
        警告列表；实验价值不足不阻断工作流。
    """
    warnings: list[str] = []
    experiments = payload.get("experiments", [])
    if not isinstance(experiments, list):
        return ["NEXT_EXPERIMENTS 的 experiments 必须是数组"]
    for index, experiment in enumerate(experiments):
        if not isinstance(experiment, dict):
            warnings.append(f"experiments[{index}] 不是对象")
            continue
        decision = experiment.get("decision")
        if not isinstance(decision, str) or not decision.strip():
            warnings.append(f"experiments[{index}] 未说明会改变哪项路线、模型或结论决定")
    return warnings


def write_next_experiments(run_dir: Path, payload: dict[str, Any]) -> dict[str, Any]:
    """受控保存只包含决策价值实验的下一实验队列。

    Args:
        run_dir: 当前运行目录。
        payload: 包含 ``experiments`` 数组的实验计划。

    Returns:
        已写入的规范对象。

    Raises:
        ContractError: 计划结构非法，或首版截止后尝试新增实验。
    """
    warnings = validate_next_experiments(payload)
    experiments = payload.get("experiments")
    if not isinstance(experiments, list):
        raise ContractError("NEXT_EXPERIMENTS 的 experiments 必须是数组")
    path = run_dir / NEXT_EXPERIMENTS_PATH
    old_ids: set[str] = set()
    if path.is_file():
        existing = load_json(path)
        old_ids = {
            item.get("experiment_id")
            for item in existing.get("experiments", [])
            if isinstance(item, dict) and isinstance(item.get("experiment_id"), str)
        }
    new_ids = {
        item.get("experiment_id")
        for item in experiments
        if isinstance(item, dict) and isinstance(item.get("experiment_id"), str)
    }
    added_ids = new_ids - old_ids
    if added_ids:
        from shumozizi.simple.delivery import require_delivery_action_allowed

        findings = [
            item.get("revision_case", str(item.get("review_finding", "")))
            for item in experiments
            if isinstance(item, dict) and item.get("experiment_id") in added_ids
        ]
        require_delivery_action_allowed(
            run_dir, "add_experiment_plan", review_findings=findings
        )
    document = {
        "schema_version": "1.0",
        "run_id": run_dir.name,
        "experiments": experiments,
        "warnings": warnings,
        "updated_at": utc_now(),
    }
    atomic_json(path, document)
    return document


def verify_submission_exports(
    run_dir: Path, answer_map: dict[str, Any] | None = None
) -> dict[str, Any]:
    """复核提交工作簿的来源结果与核心指标单元格。

    Args:
        run_dir: 当前运行目录。
        answer_map: 可选的待写答案映射；省略时读取正式答案映射。

    Returns:
        含 ``success``、``errors`` 和已复核导出记录的机械检查结果。
    """
    if answer_map is None:
        path = run_dir / ANSWER_MAP_PATH
        if not path.is_file():
            return {"success": True, "skipped": True, "errors": [], "exports": []}
        answer_map = load_json(path)
    mapping = answer_map.get("answers", answer_map)
    if not isinstance(mapping, dict):
        return {
            "success": False,
            "errors": ["answer_map 必须包含 answers 对象"],
            "exports": [],
        }
    results = {
        item["result_id"]: item for item in read_result_index(run_dir)["results"]
    }
    errors: list[str] = []
    verified: list[dict[str, Any]] = []
    for question_id, raw in mapping.items():
        if not isinstance(raw, dict):
            continue
        export = raw.get("submission_export")
        excel_location = raw.get("excel_output_location")
        if excel_location is not None and not isinstance(export, dict):
            errors.append(
                f"{question_id}.excel_output_location 必须配套 submission_export，"
                "声明 source_result_id 与 metric_cells"
            )
            continue
        if export is None:
            continue
        if not isinstance(export, dict):
            errors.append(f"{question_id}.submission_export 必须是对象")
            continue
        try:
            relative_path = _require_text(
                export.get("path"), f"{question_id}.submission_export.path"
            )
            source_result_id = _require_text(
                export.get("source_result_id"),
                f"{question_id}.submission_export.source_result_id",
            )
            metric_cells = export.get("metric_cells")
            if not isinstance(metric_cells, dict) or not metric_cells:
                raise ContractError(
                    f"{question_id}.submission_export.metric_cells 必须是非空对象"
                )
            objective_answer = raw.get("objective_answer")
            if not isinstance(objective_answer, dict):
                raise ContractError(
                    f"{question_id}.submission_export 缺少 objective_answer 绑定"
                )
            if source_result_id != objective_answer.get("result_id"):
                raise ContractError(
                    f"{question_id}.submission_export.source_result_id 必须等于 "
                    f"objective_answer {objective_answer.get('result_id')}"
                )
            if excel_location is not None and relative_path != excel_location:
                raise ContractError(
                    f"{question_id}.submission_export.path 必须与 excel_output_location 一致"
                )
            result = results.get(source_result_id)
            if (
                result is None
                or result.get("question_id") != question_id
                or result.get("execution_mode") != "production"
                or result.get("execution_valid") is not True
                or result.get("status") != "current"
                or result.get("scientific_status", "valid") == "invalidated"
            ):
                raise ContractError(
                    f"{question_id}.submission_export.source_result_id "
                    "必须是本问有效 production 结果"
                )
            workbook_path = resolve_inside(run_dir, relative_path, must_exist=True)
            if workbook_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
                raise ContractError(
                    f"{question_id}.submission_export.path 必须是 xlsx 或 xlsm 工作簿"
                )
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise ContractError(
                    "复核 Excel 提交产物需要安装 openpyxl"
                ) from exc
            try:
                workbook = load_workbook(
                    workbook_path,
                    read_only=True,
                    data_only=True,
                )
            except Exception as exc:
                # 第三方解析器会针对损坏 ZIP、旧格式和加密文件抛出不同异常；
                # 在机械 QA 边界统一转换为可定位的合同失败，不能让终检直接崩溃。
                raise ContractError(
                    f"{question_id}.submission_export 无法读取工作簿 {relative_path}: {exc}"
                ) from exc
            try:
                checked_cells: dict[str, dict[str, Any]] = {}
                for metric, reference in metric_cells.items():
                    metric_name = _require_text(
                        metric, f"{question_id}.submission_export.metric_cells metric"
                    )
                    cell_reference = _require_text(
                        reference,
                        f"{question_id}.submission_export.metric_cells.{metric_name}",
                    )
                    sheet_name, separator, coordinate = cell_reference.rpartition("!")
                    if not separator or not sheet_name or not coordinate:
                        raise ContractError(
                            f"{question_id}.{metric_name} 单元格必须使用 Sheet!A1 格式"
                        )
                    normalized_sheet = sheet_name.strip("'").replace("''", "'")
                    if normalized_sheet not in workbook.sheetnames:
                        raise ContractError(
                            f"{question_id}.{metric_name} 引用不存在的工作表 {normalized_sheet}"
                        )
                    expected = _number(
                        result.get("metrics", {}).get(metric_name),
                        f"{question_id}.{source_result_id}.{metric_name}",
                    )
                    actual = _number(
                        workbook[normalized_sheet][coordinate].value,
                        f"{question_id}.{cell_reference}",
                    )
                    tolerance = max(1e-9, abs(expected) * 1e-9)
                    if abs(actual - expected) > tolerance:
                        raise ContractError(
                            f"{question_id}.{metric_name} 工作簿值 {actual:g} 与 "
                            f"objective result {source_result_id} 的 {expected:g} 不一致"
                        )
                    checked_cells[metric_name] = {
                        "cell": cell_reference,
                        "value": actual,
                    }
            finally:
                workbook.close()
            verified.append(
                {
                    "question_id": question_id,
                    "path": relative_path,
                    "source_result_id": source_result_id,
                    "metric_cells": checked_cells,
                }
            )
        except (ContractError, OSError, ValueError) as exc:
            errors.append(str(exc))
    return {"success": not errors, "errors": errors, "exports": verified}


def write_answer_map(run_dir: Path, payload: dict[str, Any]) -> dict[str, Any]:
    """保存逐问直接答案映射，不把它提升为贡献主张合同。

    Args:
        run_dir: 当前运行目录。
        payload: 以问题 ID 为键的答案映射。

    Returns:
        已写入的规范对象。
    """
    state = read_simple_state(run_dir)
    questions = set(state["required_questions"])
    mapping = payload.get("answers", payload)
    if not isinstance(mapping, dict):
        raise ContractError("answer_map 必须是问题 ID 到答案位置的对象")
    missing = sorted(questions - set(mapping))
    if missing:
        raise ContractError("answer_map 缺少必答问题: " + ", ".join(missing))
    from shumozizi.simple.modeling_units import question_outcome_selections

    outcomes = question_outcome_selections(run_dir)
    for question_id in questions:
        item = mapping[question_id]
        if not isinstance(item, dict):
            raise ContractError(f"{question_id} 的 answer_map 条目必须是对象")
        result_ids = item.get("result_ids")
        location = item.get("direct_answer_location")
        if not isinstance(result_ids, list) or not result_ids or not all(isinstance(value, str) for value in result_ids):
            raise ContractError(f"{question_id} 必须绑定至少一个 result_id")
        _require_text(location, f"{question_id}.direct_answer_location")
        outcome = outcomes.get(question_id)
        if outcome is not None:
            objective_answer = outcome.get("objective_answer")
            if not isinstance(objective_answer, dict):
                raise ContractError(
                    f"{question_id} 尚无满足题面 endpoint、exact 指标、硬约束与"
                    "可行性的 objective_answer，必须返回分析/实验"
                )
            if objective_answer.get("claim_level") == "legacy_unverified":
                raise ContractError(
                    f"{question_id} 使用旧 oracle-only 且缺少真实 agreement；"
                    "必须迁移 MODELING_UNITS 到 1.4 后才能生成正式答案映射"
                )
            primary_result_id = _require_text(
                item.get("primary_result_id"), f"{question_id}.primary_result_id"
            )
            if primary_result_id != objective_answer["result_id"]:
                raise ContractError(
                    f"{question_id}.primary_result_id 必须等于题面原目标下的 "
                    f"objective_answer {objective_answer['result_id']}，"
                    "不得由稳健建议替换"
                )
            if primary_result_id not in result_ids:
                raise ContractError(f"{question_id}.primary_result_id 必须列入 result_ids")
            item["objective_answer"] = dict(objective_answer)
            item["recommended_plan"] = outcome.get("recommended_plan")
            item["evidence_grade"] = dict(outcome["evidence_grade"])
            item["warnings"] = list(outcome.get("warnings", []))
        excel_location = item.get("excel_output_location")
        if excel_location is not None:
            _require_text(excel_location, f"{question_id}.excel_output_location")
        insight_ids = item.get("insight_ids")
        if insight_ids is not None and (
            not isinstance(insight_ids, list)
            or not all(isinstance(value, str) and value.strip() for value in insight_ids)
        ):
            raise ContractError(f"{question_id}.insight_ids 必须是非空文本数组")
    document = {
        "schema_version": "1.1" if outcomes else "1.0",
        "run_id": run_dir.name,
        "answers": mapping,
        "generated_at": utc_now(),
    }
    export_check = verify_submission_exports(run_dir, document)
    if not export_check["success"]:
        raise ContractError("; ".join(export_check["errors"]))
    atomic_json(run_dir / ANSWER_MAP_PATH, document)
    return document
